import asyncio
import os
import io
import threading
from flask import Flask
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
from openpyxl import load_workbook, Workbook
from pymongo import MongoClient

# ================= RENDER SERVER =================
app = Flask(__name__)
@app.route('/')
def home(): return "Bot is alive!"
def run_flask(): app.run(host='0.0.0.0', port=int(os.environ.get("PORT", 10000)))

# ================= CONFIG =================
TOKEN = os.environ.get("BOT_TOKEN", "8959503198:AAGLQg8tzCE5ErVHoiGam8I2Srh75S1c8QY")
ADMIN_ID = 2106634618
MONGO_URI = os.environ.get("MONGO_URI", "")  # MongoDB Atlas URI

bot = Bot(token=TOKEN)
dp = Dispatcher(storage=MemoryStorage())

# ================= MONGODB SETUP =================
client = MongoClient(MONGO_URI)
mongo_db = client["fb_id_store"]

users_col    = mongo_db["users"]     # balance
settings_col = mongo_db["settings"]  # price
stock_col    = mongo_db["stock"]     # ID stock

# ================= DATABASE FUNCTIONS =================
def get_balance(user_id):
    user = users_col.find_one({"_id": str(user_id)})
    return int(user["balance"]) if user else 0

def add_balance(user_id, amount):
    users_col.update_one(
        {"_id": str(user_id)},
        {"$inc": {"balance": int(amount)}},
        upsert=True
    )

def get_price(category):
    doc = settings_col.find_one({"_id": f"price_{category}"})
    return int(doc["value"]) if doc else 15

def set_price(category, price):
    settings_col.update_one(
        {"_id": f"price_{category}"},
        {"$set": {"value": int(price)}},
        upsert=True
    )

# ================= STOCK FUNCTIONS =================
def get_stock_count(cat):
    return stock_col.count_documents({"category": cat})

def add_stock_from_excel(cat, file_bytes):
    wb = load_workbook(filename=io.BytesIO(file_bytes))
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        if row and any(cell is not None for cell in row):
            rows.append({
                "category": cat,
                "uid":      str(row[0]).strip().replace(".0", "") if row[0] else "N/A",
                "password": str(row[1]).strip() if len(row) > 1 and row[1] else "N/A",
                "cookies":  str(row[2]).strip() if len(row) > 2 and row[2] else "N/A",
            })
    if rows:
        stock_col.delete_many({"category": cat})
        stock_col.insert_many(rows)
    return len(rows)

def pop_stock(cat, amount):
    items = list(stock_col.find({"category": cat}).limit(amount))
    if len(items) < amount:
        return None
    ids = [item["_id"] for item in items]
    stock_col.delete_many({"_id": {"$in": ids}})
    return items

# ================= STATES =================
class BuyState(StatesGroup):
    category = State()
    amount   = State()

class DepositState(StatesGroup):
    method = State()
    amount = State()
    txnid  = State()

class AdminState(StatesGroup):
    upload_cat = State()
    price_cat  = State()
    new_price  = State()

# ================= KEYBOARDS =================
def main_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🛍️ BUY FACEBOOK ID", callback_data="products")],
        [InlineKeyboardButton(text="💳 Deposit",    callback_data="deposit"),
         InlineKeyboardButton(text="💰 My Balance", callback_data="balance")],
        [InlineKeyboardButton(text="🧑‍💻 Live Support", url="https://t.me/mohammadrony56"),
         InlineKeyboardButton(text="👨‍💻 Admin",        callback_data="admin")]
    ])

def admin_panel():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📂 Upload Stock",      callback_data="admin_upload")],
        [InlineKeyboardButton(text="💰 Set ID Price",      callback_data="admin_price")],
        [InlineKeyboardButton(text="💡 Balance Guide",     callback_data="balhelp")],
        [InlineKeyboardButton(text="🔙 Back to Main Menu", callback_data="back")]
    ])

def deposit_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📱 bKash", callback_data="dep_bkash"),
         InlineKeyboardButton(text="📱 Nagad", callback_data="dep_nagad")],
        [InlineKeyboardButton(text="🔙 Main Menu", callback_data="back")]
    ])

def back_btn():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Main Menu", callback_data="back")]
    ])

# ================= DESIGN =================
def get_dashboard_text(name, user_id):
    return (
        f"▬▬▬ ✦ PREMIUM ID STORE ✦ ▬▬▬\n\n"
        f"👤 **User::** {name}\n"
        f"🆔 **User ID::** `{user_id}`\n\n"
        f"──────────────────────────────\n"
        f"👋 **আমাদের বটে আপনাকে স্বাগতম!**\n"
        f"📥 স্টক শেষ হওয়ার আগেই নিচের বাটনে ক্লিক করে আপনার আইডি সংগ্রহ করুন 🎯!"
    )

# ================= HANDLERS =================
@dp.message(Command("start"))
async def start(message: types.Message):
    await message.answer(
        get_dashboard_text(message.from_user.full_name, message.from_user.id),
        reply_markup=main_menu(), parse_mode="Markdown"
    )

@dp.message(Command("addbal"))
async def addbal(message: types.Message):
    if message.from_user.id != ADMIN_ID: return
    try:
        _, uid, amount = message.text.split()
        add_balance(uid, int(amount))
        await message.answer(f"✅ **Manual Success!** `{amount} BDT` added to `{uid}`.", parse_mode="Markdown")
    except:
        await message.answer("❌ Use: `/addbal user_id amount`", parse_mode="Markdown")

@dp.message(lambda m: m.document)
async def upload_stock(message: types.Message, state: FSMContext):
    if message.from_user.id != ADMIN_ID: return
    current_state = await state.get_state()
    if current_state != AdminState.upload_cat:
        await message.answer("⚠️ আগে এডমিন প্যানেল থেকে সিলেক্ট করুন কোন ক্যাটাগরির স্টক আপলোড করতে চান।")
        return
    user_data = await state.get_data()
    cat = user_data["upload_cat"]
    try:
        file = await bot.get_file(message.document.file_id)
        file_bytes = await bot.download_file(file.file_path)
        count = add_stock_from_excel(cat, file_bytes.read())
        await message.answer(
            f"🚀 **{cat} PC CLONE Stock Updated!**\n✅ মোট `{count}` টি ID যোগ হয়েছে।",
            reply_markup=admin_panel(), parse_mode="Markdown"
        )
        await state.clear()
    except Exception as e:
        await message.answer(f"❌ **Upload Failed!** Error: `{e}`", reply_markup=admin_panel(), parse_mode="Markdown")
        await state.clear()

# ================= CALLBACKS =================
@dp.callback_query()
async def callback(call: types.CallbackQuery, state: FSMContext):
    data    = call.data
    user_id = str(call.from_user.id)
    name    = call.from_user.full_name
    try:
        await call.answer()
    except: pass

    if data == "balance":
        await call.message.edit_text(
            f"▬▬▬ ✦ ACCOUNT BALANCE ✦ ▬▬▬\n\n"
            f"👤 User: **{name}**\n"
            f"💰 Available Balance: `{get_balance(user_id)} BDT` 💵\n\n"
            f"──────────────────────────────\n"
            f"ℹ️ _Need more balance? Click on Deposit to add funds._",
            reply_markup=back_btn(), parse_mode="Markdown"
        )

    elif data == "products":
        await call.message.edit_text(
            f"📦 **PRODUCT CATALOGUE**\n\n"
            f"🔥 **1. 1000x PC CLONE ID**\n"
            f"├ Stock: `{get_stock_count('1000x')} Pcs`\n"
            f"└ Price: `{get_price('1000x')} BDT` / Pcs\n\n"
            f"🔥 **2. 61x PC CLONE ID**\n"
            f"├ Stock: `{get_stock_count('61x')} Pcs`\n"
            f"└ Price: `{get_price('61x')} BDT` / Pcs\n\n"
            f"👇 আপনি কোন ক্যাটাগরির আইডি কিনতে চান তা নিচে সিলেক্ট করুন:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🛒 Buy 1000x PC CLONE", callback_data="buy_1000x")],
                [InlineKeyboardButton(text="🛒 Buy 61x PC CLONE",   callback_data="buy_61x")],
                [InlineKeyboardButton(text="🔙 Main Menu",           callback_data="back")]
            ]), parse_mode="Markdown"
        )

    elif data.startswith("buy_"):
        cat = data.split("_")[1]
        await state.update_data(category=cat)
        await state.set_state(BuyState.amount)
        await call.message.answer(
            f"🔢 **আপনি কত পিস {cat} PC CLONE আইডি কিনতে চান?**\n\n"
            f"👉 দয়া করে সংখ্যাটি লিখে পাঠান:", parse_mode="Markdown"
        )

    elif data == "deposit":
        await call.message.edit_text(
            "📥 **CHOOSE DEPOSIT METHOD**\n\nনিচের যেকোনো একটি পেমেন্ট গেটওয়ে সিলেক্ট করুন:",
            reply_markup=deposit_menu(), parse_mode="Markdown"
        )

    elif data.startswith("dep_"):
        method = "bKash" if "bkash" in data else "Nagad"
        await state.update_data(method=method)
        await state.set_state(DepositState.amount)
        await call.message.edit_text(
            f"📱 **{method} (Personal):** `01639216168`\n"
            f"📢 **সর্বনিম্ন ডিপোজিট:** `20` টাকা।\n\n"
            f"💵 **ধাপ ১:** প্রথমে ওপরের নাম্বারে **Send Money** করুন।\n"
            f"📥 **ধাপ ২:** কত টাকা পাঠিয়েছেন তা লিখুন (যেমন: `150`):",
            parse_mode="Markdown"
        )

    elif data == "admin":
        if call.from_user.id != ADMIN_ID:
            await call.answer("❌ দুঃখিত! আপনি অ্যাডমিন নন।", show_alert=True)
            return
        try:
            await call.message.edit_text("⚙️ **ADMIN CONTROL PANEL**", reply_markup=admin_panel(), parse_mode="Markdown")
        except Exception as e:
            print(f"Error: {e}")

    elif data == "admin_upload":
        await call.message.edit_text(
            "📂 **কোন ক্যাটাগরির স্টক ফাইল আপলোড করতে চান?**",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📁 Upload 1000x Stock", callback_data="upcat_1000x")],
                [InlineKeyboardButton(text="📁 Upload 61x Stock",   callback_data="upcat_61x")],
                [InlineKeyboardButton(text="🔙 Admin Panel",        callback_data="admin")]
            ])
        )

    elif data.startswith("upcat_"):
        cat = data.split("_")[1]
        await state.update_data(upload_cat=cat)
        await state.set_state(AdminState.upload_cat)
        await call.message.edit_text(
            f"📂 **Please send the `.xlsx` file for {cat} PC CLONE now...**", parse_mode="Markdown"
        )

    elif data == "admin_price":
        await call.message.edit_text(
            "💰 **কোন ক্যাটাগরির দাম পরিবর্তন করতে চান?**",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="💰 Set 1000x Price", callback_data="prcat_1000x")],
                [InlineKeyboardButton(text="💰 Set 61x Price",   callback_data="prcat_61x")],
                [InlineKeyboardButton(text="🔙 Admin Panel",     callback_data="admin")]
            ])
        )

    elif data.startswith("prcat_"):
        cat = data.split("_")[1]
        await state.update_data(price_cat=cat)
        await state.set_state(AdminState.new_price)
        await call.message.edit_text(
            f"💰 **SET {cat.upper()} PRICE**\n\n📊 বর্তমান দাম: `{get_price(cat)} BDT`\n\n👉 নতুন দামটি লিখুন:",
            parse_mode="Markdown"
        )

    elif data == "balhelp":
        await call.message.edit_text("💡 `/addbal <user_id> <amount>`", reply_markup=admin_panel(), parse_mode="Markdown")

    elif data == "back":
        await state.clear()
        await call.message.edit_text(get_dashboard_text(name, user_id), reply_markup=main_menu(), parse_mode="Markdown")

    elif data.startswith("approve_"):
        parts = data.split("_", 2); target_uid, amount = parts[1], parts[2]
        add_balance(target_uid, amount)
        await call.message.edit_text(f"✅ **Deposit Approved!**\nUser `{target_uid}` কে `{amount} BDT` দেওয়া হয়েছে।", parse_mode="Markdown")
        try:
            await bot.send_message(target_uid, f"🎉 **Deposit Approved!**\n\n`{amount} BDT` has been credited to your account.", parse_mode="Markdown")
        except: pass

    elif data.startswith("reject_"):
        parts = data.split("_", 1); target_uid = parts[1]
        await call.message.edit_text("❌ **Deposit Rejected!**", parse_mode="Markdown")
        try:
            await bot.send_message(target_uid, "❌ **Deposit Rejected!**\n\nYour transaction was invalid.", parse_mode="Markdown")
        except: pass

# ================= ADMIN PRICE SET =================
@dp.message(AdminState.new_price)
async def process_new_price(message: types.Message, state: FSMContext):
    if message.from_user.id != ADMIN_ID: return
    try:
        new_price = int(message.text)
        if new_price <= 0: raise ValueError
    except:
        await message.answer("❌ **ভুল ইনপুট!** শুধু সংখ্যা লিখুন।")
        return
    user_data = await state.get_data()
    cat = user_data["price_cat"]
    set_price(cat, new_price)
    await message.answer(
        f"✅ **সফল!** {cat} এর নতুন দাম `{new_price} BDT` সেট হয়েছে।",
        reply_markup=admin_panel(), parse_mode="Markdown"
    )
    await state.clear()

# ================= DEPOSIT PROCESS =================
MIN_DEPOSIT = 20

@dp.message(DepositState.amount)
async def dep_amount(message: types.Message, state: FSMContext):
    try:
        amount = int(message.text)
        if amount <= 0: raise ValueError
    except:
        await message.answer("❌ **ভুল ইনপুট!** সঠিক টাকার পরিমাণ সংখ্যায় লিখুন।")
        return
    if amount < MIN_DEPOSIT:
        await message.answer(f"⚠️ সর্বনিম্ন ডিপোজিট `{MIN_DEPOSIT} BDT`। আবার সঠিক পরিমাণ লিখুন:")
        return
    await state.update_data(amount=amount)
    await state.set_state(DepositState.txnid)
    await message.answer("🆔 **ধাপ ৩:** পেমেন্টের **Transaction ID (TxnID)** লিখে পাঠান:")

@dp.message(DepositState.txnid)
async def dep_txnid(message: types.Message, state: FSMContext):
    txnid     = message.text.strip()
    user_data = await state.get_data()
    method    = user_data['method']
    amount    = user_data['amount']
    user_id   = message.from_user.id

    await message.answer("⏳ **পেমেন্ট রিকোয়েস্ট সাবমিট হয়েছে!** এডমিন ভেরিফাই করার পর ব্যালেন্স এড হবে।", reply_markup=back_btn())
    await bot.send_message(
        ADMIN_ID,
        f"📥 **NEW DEPOSIT REQUEST**\n\n👤 **User:** {message.from_user.full_name} (`{user_id}`)\n💳 **Method:** {method}\n💰 **Amount:** `{amount} BDT`\n🔍 **TxnID:** `{txnid}`",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="✅ Approve", callback_data=f"approve_{user_id}_{amount}"),
            InlineKeyboardButton(text="❌ Reject",  callback_data=f"reject_{user_id}")
        ]]),
        parse_mode="Markdown"
    )
    await state.clear()

# ================= BUY PROCESS =================
@dp.message(BuyState.amount)
async def buy_amount(message: types.Message, state: FSMContext):
    user_id = str(message.from_user.id)
    try:
        amount = int(message.text)
        if amount <= 0: raise ValueError
    except:
        await message.answer("❌ **Invalid Input!** Please send a valid positive number.")
        return

    user_data   = await state.get_data()
    cat         = user_data["category"]
    stock_count = get_stock_count(cat)

    if stock_count < amount:
        await message.answer(f"❌ **Insufficient Stock!** Only `{stock_count}` IDs available in {cat}.", parse_mode="Markdown")
        await state.clear()
        return

    price_per_id = get_price(cat)
    total_price  = amount * price_per_id

    if get_balance(user_id) < total_price:
        await message.answer(
            f"❌ **Insufficient Balance!** Need `{total_price} BDT`.\n\n👉 ব্যালেন্স চেক করতে 'My Balance' বাটনে ক্লিক করুন।",
            parse_mode="Markdown"
        )
        await state.clear()
        return

    bought = pop_stock(cat, amount)
    if not bought:
        await message.answer("❌ Stock error! আবার চেষ্টা করুন।")
        await state.clear()
        return

    add_balance(user_id, -total_price)

    # Disk ছাড়া memory-তে Excel তৈরি
    wb = Workbook()
    ws = wb.active
    ws.append(["UID", "PASSWORD", "COOKIES"])
    for item in bought:
        ws.append([item.get("uid", "N/A"), item.get("password", "N/A"), item.get("cookies", "N/A")])

    file_buffer = io.BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)

    try:
        await message.answer_document(
            BufferedInputFile(file_buffer.read(), filename=f"{cat}_ids.xlsx"),
            caption=(
                f"🎉 **ORDER DELIVERED!**\n\n"
                f"📦 **Category:** `{cat} PC CLONE`\n"
                f"🔢 **Items:** `{amount} Pcs`\n"
                f"💰 **Debited:** `{total_price} BDT`"
            ),
            parse_mode="Markdown"
        )
    except Exception as e:
        await message.answer(f"❌ Delivery Error: `{e}`")
    await state.clear()

# ================= RUN =================
async def main():
    threading.Thread(target=run_flask).start()
    print("🔥 Bot Running with MongoDB...")
    await dp.start_polling(bot, skip_updates=True, drop_pending_updates=True)

if __name__ == "__main__":
    asyncio.run(main())
